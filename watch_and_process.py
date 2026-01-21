# watch_and_process.py
# --------------------
# "Process SR Report" — Local PC watcher for TeamBinder reports
# - watches Downloads (or a user-selected folder)
# - only processes files whose names start with "ETSA-TSA-" and end with ".xls"
# - file must be no older than 12 hours
# - converts .xls -> .xlsx (Excel COM) if needed
# - transforms Column B (from row 19) by appending " ### <url>" once
# - copies processed file into a user-selected SharePoint synced folder (OneDrive uploads)
#
# First-run wizard:
#  - explains that the folder MUST be synced locally via OneDrive (SharePoint Sync)
#  - prompts user to select the local SharePoint library folder
#  - optionally lets user pick a different watch folder instead of Downloads
#
# Usage:
#   - Double-click the EXE to start the watcher (manual start; no startup install)
#   - --reset  : re-run the folder selection wizard
#
# Author: "Process SR Report" build
# Date: Jan 2026

import os, re, sys, time, json, shutil, logging
from pathlib import Path
from datetime import datetime, timedelta

from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

try:
    import win32com.client as win32  # Excel COM
except Exception:
    win32 = None

from openpyxl import load_workbook

import tkinter as tk
from tkinter import messagebox, filedialog

APP_DISPLAY_NAME = "Process SR Report"
APP_DIR_NAME     = "ProcessSRReport"  # for %APPDATA% folder and log
DEFAULT_WATCH_GLOB = "ETSA-TSA-*.xls"  # strict .xls as per requirement
DEFAULT_START_ROW  = 19
DEFAULT_COL_B      = 2
DEFAULT_SEPARATOR  = " ### "
DEFAULT_TS_FMT     = "%Y%m%d_%H%M%S"
MAX_FILE_AGE_HOURS = 12


def get_appdata_dir() -> Path:
    base = os.getenv("APPDATA") or str(Path.home() / "AppData" / "Roaming")
    p = Path(base) / APP_DIR_NAME
    p.mkdir(parents=True, exist_ok=True)
    return p

SETTINGS_PATH = get_appdata_dir() / "settings.json"
LOG_PATH      = get_appdata_dir() / "bot.log"


def load_settings():
    if SETTINGS_PATH.exists():
        try:
            return json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_settings(obj: dict):
    SETTINGS_PATH.write_text(json.dumps(obj, indent=2), encoding="utf-8")


def default_downloads():
    return str(Path.home() / "Downloads")


def prompt_explain_and_pick_sharepoint():
    root = tk.Tk(); root.withdraw()
    message = (
        "Select your local SharePoint library folder.\n\n"
        "Important: The library must be synced to your PC via OneDrive first.\n"
        "In SharePoint, open the library and click the 'Sync' button. Once it "
        "appears in File Explorer under your organization, pick that folder here."
    )
    messagebox.showinfo(title=f"{APP_DISPLAY_NAME} — SharePoint folder required", message=message)
    sp_path = filedialog.askdirectory(title="Select your synced SharePoint library folder")
    if not sp_path:
        messagebox.showwarning("Setup incomplete", "No folder selected. Exiting.")
        sys.exit(0)
    return sp_path


def prompt_optional_watch_folder(default_dl):
    root = tk.Tk(); root.withdraw()
    if messagebox.askyesno(
        f"{APP_DISPLAY_NAME} — Watch folder",
        f"Default watch folder is your Downloads:\n\n{default_dl}\n\n"
        "Do you want to choose a different folder?"
    ):
        path = filedialog.askdirectory(title="Select the folder to watch for TeamBinder reports")
        return path or default_dl
    return default_dl


def ensure_config(reset=False):
    cfg = load_settings()
    if reset:
        cfg = {}
    changed = False

    if not cfg.get("sharepoint_folder") or not Path(cfg["sharepoint_folder"]).exists():
        sp = prompt_explain_and_pick_sharepoint()
        if not Path(sp).exists():
            tk.messagebox.showerror("Invalid folder", "That path does not exist. Exiting.")
            sys.exit(1)
        cfg["sharepoint_folder"] = sp
        changed = True

    if not cfg.get("watch_folder") or not Path(cfg["watch_folder"]).exists():
        dl = default_downloads()
        wf = prompt_optional_watch_folder(dl)
        if not Path(wf).exists():
            tk.messagebox.showerror("Invalid folder", "That path does not exist. Exiting.")
            sys.exit(1)
        cfg["watch_folder"] = wf
        changed = True

    if "pattern" not in cfg:
        cfg["pattern"] = DEFAULT_WATCH_GLOB
        changed = True
    if "start_row" not in cfg:
        cfg["start_row"] = DEFAULT_START_ROW
        changed = True
    if "col_b" not in cfg:
        cfg["col_b"] = DEFAULT_COL_B
        changed = True
    if "separator" not in cfg:
        cfg["separator"] = DEFAULT_SEPARATOR
        changed = True

    if changed:
        save_settings(cfg)

    return cfg


def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(sys.stdout),
                  logging.FileHandler(LOG_PATH, encoding="utf-8")]
    )
    logging.info(f"{APP_DISPLAY_NAME} started. Log: {LOG_PATH}")


def is_file_stable(p: Path, stable_secs=5) -> bool:
    last = -1; same = 0
    while same < stable_secs:
        if not p.exists():
            return False
        size = p.stat().st_size
        if size == last and size > 0:
            same += 1
        else:
            same = 0; last = size
        time.sleep(1)
    return True


def is_file_fresh_enough(p: Path, max_hours=12) -> bool:
    try:
        mtime = datetime.fromtimestamp(p.stat().st_mtime)
    except Exception:
        return False
    age = datetime.now() - mtime
    return age <= timedelta(hours=max_hours)


def excel_xls_to_xlsx(src_xls: Path, visible=False) -> Path:
    if win32 is None:
        raise RuntimeError("Excel automation not available (pywin32 not installed).")
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = bool(visible)
    wb = excel.Workbooks.Open(str(src_xls))
    dst = src_xls.with_suffix(".xlsx")
    wb.SaveAs(str(dst), FileFormat=51)  # 51 = xlOpenXMLWorkbook (.xlsx)
    wb.Close(SaveChanges=False)
    excel.Application.Quit()
    return dst


def parse_hyperlink_formula(formula: str):
    s = formula.strip()
    if not s.upper().startswith("=HYPERLINK("):
        return None, None
    inside = s[s.find("(")+1:].rstrip(")")
    inq = False; buf = ""; parts = []
    for ch in inside:
        if ch == '"':
            inq = not inq; buf += ch; continue
        if not inq and ch in ",;":
            parts.append(buf.strip()); buf = ""; continue
        buf += ch
    if buf: parts.append(buf.strip())
    def clean(t): return t[1:-1] if t.startswith('"') and t.endswith('"') else t
    url = clean(parts[0]) if parts else ""
    text = clean(parts[1]) if len(parts) > 1 else ""
    return url or None, text or None


def transform_workbook(xlsx_path: Path, start_row=19, col_b=2, sep=" ### "):
    wb = load_workbook(filename=str(xlsx_path))
    ws = wb.worksheets[0]
    max_row = ws.max_row
    if max_row < start_row:
        wb.save(str(xlsx_path))
        return 0
    changed = 0
    for r in range(start_row, max_row+1):
        cell = ws.cell(row=r, column=col_b)
        val = "" if cell.value is None else str(cell.value)
        url = None

        if cell.hyperlink and getattr(cell.hyperlink, "target", None):
            url = cell.hyperlink.target
        elif isinstance(val, str) and val.startswith("="):
            u, t = parse_hyperlink_formula(val)
            if u:
                url = u
                if t:
                    val = t
        if not url and isinstance(val, str):
            m = re.search(r"https?://\S+", val, re.IGNORECASE)
            if m: url = m.group(0)

        if url and sep not in val:
            cell.value = f"{val}{sep}{url}"
            changed += 1
        else:
            cell.value = val

    wb.save(str(xlsx_path))
    return changed


def copy_to_sharepoint(processed_path: Path, sp_folder: Path) -> Path:
    sp_folder.mkdir(parents=True, exist_ok=True)
    dest = sp_folder / processed_path.name
    shutil.copy2(processed_path, dest)
    return dest


class NewFileHandler(FileSystemEventHandler):
    def __init__(self, cfg):
        self.cfg = cfg
        self.pattern = cfg.get("pattern", DEFAULT_WATCH_GLOB)
        self.start_row = int(cfg.get("start_row", DEFAULT_START_ROW))
        self.col_b = int(cfg.get("col_b", DEFAULT_COL_B))
        self.sep = cfg.get("separator", DEFAULT_SEPARATOR)
        self.sp_folder = Path(cfg["sharepoint_folder"]).resolve()

    def _matches(self, name: str) -> bool:
        return Path(name).match(self.pattern)

    def _process(self, path_str):
        try:
            p = Path(path_str)
            if p.is_dir(): return
            if not self._matches(p.name): return

            # Age check first
            if not is_file_fresh_enough(p, MAX_FILE_AGE_HOURS):
                logging.info(f"Skipping old file (> {MAX_FILE_AGE_HOURS}h): {p}")
                return

            logging.info(f"Detected candidate: {p}")
            if not is_file_stable(p, 5):
                logging.warning(f"File not stable: {p}")
                return

            src = p
            # Convert .xls -> .xlsx if needed
            if src.suffix.lower() == ".xls":
                if win32 is None:
                    logging.error("Received .xls but Excel automation unavailable; skipping.")
                    return
                logging.info(f"Converting .xls to .xlsx using Excel: {src}")
                src = excel_xls_to_xlsx(src, visible=False)
                logging.info(f"Converted to: {src}")

            if src.suffix.lower() != ".xlsx":
                logging.warning(f"Unsupported extension {src.suffix}; skipping.")
                return

            logging.info(f"Transforming workbook: {src}")
            changed = transform_workbook(src, self.start_row, self.col_b, self.sep)
            logging.info(f"Transform complete; rows changed: {changed}")

            stamp = datetime.now().strftime(DEFAULT_TS_FMT)
            processed_name = f"Processed_{stamp}.xlsx"
            processed_path = src.with_name(processed_name)
            shutil.copy2(src, processed_path)

            dest = copy_to_sharepoint(processed_path, self.sp_folder)
            logging.info(f"Copied to SharePoint folder (OneDrive will sync): {dest}")

        except Exception as e:
            logging.exception(f"Error processing file: {path_str}: {e}")

    def on_created(self, event):  self._process(event.src_path)
    def on_modified(self, event): self._process(event.src_path)


def main():
    reset = "--reset" in [a.lower() for a in sys.argv[1:]]
    cfg = ensure_config(reset=reset)
    setup_logging()
    watch_folder = Path(cfg["watch_folder"]).resolve()
    pattern = cfg["pattern"]
    logging.info(f"Watching: {watch_folder} | Pattern: {pattern}")
    logging.info(f"Output SharePoint folder: {cfg['sharepoint_folder']}")

    handler = NewFileHandler(cfg)
    observer = Observer()
    observer.schedule(handler, str(watch_folder), recursive=False)
    observer.start()
    logging.info("Watcher is active. Close this window to stop.")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        logging.info("Stopping watcher...")
        observer.stop()
    observer.join()


if __name__ == "__main__":
    main()
